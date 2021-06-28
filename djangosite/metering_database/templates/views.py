from django.shortcuts import render
from django.http import HttpResponse, request
import datetime
from datetime import datetime,date
import time
import pymysql
import pandas as pd
import openpyxl as py
from pygrok import Grok
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import AnonymousUser
import openpyxl as xl
#import datetime
import xlsxwriter
from sqlalchemy import create_engine
from sqlalchemy.types import String, SmallInteger,VARCHAR,INT
import matplotlib.pyplot as plt
import mpld3
from django.core.files.storage import FileSystemStorage
import io
import calendar
from openpyxl import Workbook
import csv
import os
conn = mysql.connector.connect(**db_config)
c = conn.cursor()


def index(request):
    return render(request, 'index.html')

def homepage(request):
    #user = user_act
    #user = active_user
    return render(request, 'index.html')

def welcome(request):
    if request.method == 'POST':
        user_act = request.POST['username']
        password_given = request.POST['password']
        user = authenticate(username = user_act, password = password_given)
        #if user.is_authenticated == True:
        if user is not None:
            global active_user
            active_user = user_act
            AnonymousUser = user
            return render(request, 'welcome.html',{'user':user})

        else:
            error = 'Please check your username and password'
            return render(request, 'index.html', {'error':error})


def meterrecord(request):  #inserting meter records for standalone meter
    if request.method == 'POST':
        unix = datetime.now().replace (microsecond=0)
        time_added = unix
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        village = request.POST['village_st']
        subcounty = request.POST['sub_county_st']
        district = request.POST['district_st']
        region = request.POST['region_st']
        x_cordinates = request.POST['x_cordinates_st']
        y_cordinates = request.POST['y_cordinates_st']
        feeder_name = request.POST['feeder_name_st']
        voltage = request.POST['voltage_st']

        distributor_1 = request.POST['distributor_1']
        distributor_2 = request.POST['distributor_2']
        meter_owner = request.POST['meter_owner_st']


        manu_st = request.POST['manu_meter_st'] #second take
        unit_type = request.POST['meter_type_st']
        D_O_M = request.POST['DOM_st']
        MU_Serial_No = request.POST['mu_meter_no_st']
        no_of_elements = request.POST['elements_st']
        wired_as = request.POST['wired_as_st']
        comm_date = request.POST['comm_date_st']
        DLP = request.POST['DLP_st']
        metering_cores = request.POST['meter_cores_st']
        CTratios = request.POST['core_ratios_st']
        core_used = request.POST['core_used_st']
        core_used_accuracy = request.POST['accuracy_st']
        avail_spares = request.POST['avail_spares_st']
        spares_class = request.POST['spares_class_st']
        VTratio = request.POST['vt_ratio_st']
        VT_accuracy = request.POST['vt_accuracy_st']
        #meter_no = request.POST['meter_no_st']
        #global meter_no
        #meter_no = request.POST['meter_no']
        global active_user
        added_by = active_user
        #global meter

        meter_manuf = request.POST['energy_meter_manu'] #main energy meter details
        meter_type = request.POST['energy_meter_type']
        meter_no = request.POST['energy_meter_no']
        Y_O_M = request.POST['meter_YOM']
        meter_accuracy = request.POST['e_meter_accuracy']
        no_of_meter_elements = request.POST['e_meter_elements']
        meter_wired_as = request.POST['e_meter_wired']
        meter_install_date = request.POST['e_instal_date']
        meter_decom_date = request.POST['e_decom_date']
        access = request.POST['access']
        IP_Address = request.POST['ip_address']
        avail_interfaces = request.POST['interfaces']
        comm_protocol = request.POST['comm_protocol']
        comm_protocol_used = request.POST['comm_protocol_used']

        meter_manuf_ch = request.POST['energy_meter_manu_ch']  # main energy meter details
        meter_type_ch = request.POST['energy_meter_type_ch']
        meter_no_ch = request.POST['energy_meter_no_ch']
        Y_O_M_ch = request.POST['meter_YOM_ch']
        meter_accuracy_ch = request.POST['e_meter_accuracy_ch']
        no_of_meter_elements_ch = request.POST['e_meter_elements_ch']
        meter_wired_as_ch = request.POST['e_meter_wired_ch']
        meter_install_date_ch = request.POST['e_instal_date_ch']
        meter_decom_date_ch = request.POST['e_decom_date_ch']
        access_ch = request.POST['access_ch']
        IP_Address_ch = request.POST['ip_address_ch']
        avail_interfaces_ch = request.POST['interfaces_ch']
        comm_protocol_ch = request.POST['comm_protocol_ch']
        comm_protocol_used_ch = request.POST['comm_protocol_used_ch']

        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        sql_1 = """INSERT INTO standalone_meters (time_added,added_by,village,subcounty,district,region,x_cordinates,y_cordinates,feeder_name,
         voltage,distributor_1,distributor_2,meter_owner,manufacturer,unit_type,D_O_M,MU_Serial_No,no_of_elements,wired_as,comm_date,DLP,metering_cores,CTratios,
          core_used,core_used_accuracy,avail_spares,spares_class,VTratio,VT_accuracy,meter_manuf,meter_type,meter_no,Y_O_M,meter_accuracy,no_of_meter_elements,
          meter_wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocol,comm_protocol_used,meter_manuf_ch,meter_type_ch,meter_no_ch,
          Y_O_M_ch,meter_accuracy_ch,no_of_meter_elements_ch,meter_wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocol_ch,
          comm_protocol_used_ch) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        values = (unix,added_by,village,subcounty,district,region,x_cordinates,y_cordinates,feeder_name,voltage,distributor_1,distributor_2,meter_owner,manu_st,
                  unit_type,D_O_M,MU_Serial_No,no_of_elements,wired_as,comm_date,DLP,metering_cores,CTratios,core_used,core_used_accuracy,avail_spares,spares_class,VTratio,VT_accuracy,
                  meter_manuf,meter_type,meter_no,Y_O_M,meter_accuracy,no_of_meter_elements,meter_wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,
                  comm_protocol,comm_protocol_used,meter_manuf_ch,meter_type_ch,meter_no_ch,Y_O_M_ch,meter_accuracy_ch,no_of_meter_elements_ch,meter_wired_as_ch,
                  meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocol_ch,comm_protocol_used_ch)
        c.execute(sql_1, values)
        c.close()
        conn.commit()
        conn.close()
        return render(request, 'form_success.html')

def sub_meterrecord(request):  #inserting substation meter records
    if request.method == 'POST':
        unix = datetime.now().replace (microsecond=0)
        time_added = unix

        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        substation = request.POST['sub']
        voltage = request.POST['voltage']
        district = request.POST['sub_dist']
        region = request.POST['sub_region']
        x_cordinates = request.POST['sub_cord_x']
        y_cordinates = request.POST['sub_cord_y']
        feeder_name = request.POST['sub_feeder'] #second take
        feeder_voltage = request.POST['sub_feeder_voltage']
        meter_owner = request.POST['sub_meter_owner']
        distributor = request.POST['distributor']
        contractor = request.POST['subs_contractor']
        ct_manufacturer = request.POST['sub_ct_manu'] #Third take
        ct_type = request.POST['sub_ct_type']
        no_ct_cores = request.POST['sub_ct_cores']
        ct_ratios = request.POST['sub_ct_core_ratios']
        core_used = request.POST['sub_core_used']
        accuracy_class = request.POST['sub_accuracy_class']
        avail_spares = request.POST['sub_avail_spares']
        spares_class = request.POST['sub_spares_class']
        VT_manufacturer = request.POST['sub_vt_manu']
        VT_type = request.POST['sub_VT_type']
        VT_ratio = request.POST['sub_vt_ratio']
        VT_accuracy = request.POST['sub_vt_accuracy']
        meter_manufacturer = request.POST['sub_meter_manu'] #fourth take
        meter_type = request.POST['sub_meter_type']
        meter_no = request.POST['sub_meter_no']
        meter_YOM = request.POST['sub_YOM']
        meter_accuracy_class = request.POST['sub_meter_acc_class']
        no_of_elements = request.POST['sub_meter_elements']
        wired_as = request.POST['sub_meter_wire']
        meter_install_date = request.POST['sub_meter_install_date']
        meter_decom_date = request.POST['sub_meter_decom_date']
        access = request.POST['sub_meter_access']
        IP_Address = request.POST['sub_ip_address']
        avail_interfaces = request.POST['sub_meter_interfaces']
        comm_protocols = request.POST['sub_meter_comm_protocols']
        protocol_used = request.POST['sub_meter_protocol_used']


        meter_manufacturer_ch = request.POST['sub_meter_manu_ch']  # fifth take
        meter_type_ch = request.POST['sub_meter_type_ch']
        meter_no_ch = request.POST['sub_meter_no_ch']
        meter_YOM_ch = request.POST['sub_YOM_ch']
        meter_accuracy_class_ch = request.POST['sub_meter_acc_class_ch']
        no_of_elements_ch = request.POST['sub_meter_elements_ch']
        wired_as_ch = request.POST['sub_meter_wire_ch']
        meter_install_date_ch = request.POST['sub_meter_install_date_ch']
        meter_decom_date_ch = request.POST['sub_meter_decom_date_ch']
        access_ch = request.POST['sub_meter_access_ch']
        IP_Address_ch = request.POST['sub_ip_address_ch']
        avail_interfaces_ch = request.POST['sub_meter_interfaces_ch']
        comm_protocols_ch = request.POST['sub_meter_comm_protocols_ch']
        protocol_used_ch = request.POST['sub_meter_protocol_used_ch']


        #global meter_no
        #meter_no = request.POST['meter_no']
        global active_user
        added_by = active_user
        #global meter

        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        sql_2 = """INSERT INTO substation_meters (time_added,added_by,substation,voltage, district,region,x_cordinates,y_cordinates,feeder_name,feeder_voltage,meter_owner,
                    distributor,contractor,ct_manufacturer,ct_type,no_ct_cores,ct_ratios,core_used,accuracy_class,avail_spares,spares_class,VT_manufacturer,VT_type,VT_ratio,VT_accuracy,
                    meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy_class,no_of_elements,wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocols,protocol_used,
                    meter_manufacturer_ch,meter_type_ch,meter_no_ch,meter_YOM_ch,meter_accuracy_class_ch,no_of_elements_ch,wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocols_ch,protocol_used_ch) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                     %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        values = (unix,added_by,substation,voltage,district,region,x_cordinates,y_cordinates,feeder_name,feeder_voltage,meter_owner,distributor,contractor,
                  ct_manufacturer,ct_type,no_ct_cores,ct_ratios,core_used,accuracy_class,avail_spares,spares_class,VT_manufacturer,VT_type,VT_ratio,VT_accuracy,
                  meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy_class,no_of_elements,wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocols,protocol_used,
                  meter_manufacturer_ch,meter_type_ch,meter_no_ch,meter_YOM_ch,meter_accuracy_class_ch,no_of_elements_ch,wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocols_ch,protocol_used_ch)

        c.execute(sql_2, values)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')

def new_submission(request):
    #user = user_act
    #user = active_user
    return render(request, 'welcome.html')

def success(request):
    return render(request, 'uetcl_meter.html')
    #if request.method == 'POST':
        #user = request.POST['username']
        #password_given = request.POST['password']
        #unix = int(time.time())
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
now = datetime.now()
#assigning the excel sheet a dynamic name
xlm = now.strftime('%B')
#import csv
def fillform(request):

    if request.method == 'POST':

        global sales_month,imp_wh, exp_wh, app_pow, rate1, rate2, rate3, rate4, rate5, rate6, max_dem_1, max_dem_1t, max_dem_2
        global max_dem_2t, max_dem_3, max_dem_3t, imp_mvarh, exp_mvarh, total_mvarh, resets, reset_time, pow_down_count
        global pow_down_dt, prog_count, prog_count_dt
        global total_exp_MVArh, total_imp_MVArh,sales

        listz = []
        csv_list = []
        global active_user, our_month,unix,our_year,our_file_name,df
        unix = datetime.now().time().replace(microsecond=0)
        attache = str(unix)
        pp = f'files_for_{active_user}'
        storage = str(pp)
        for f in request.FILES.getlist('filename_m'):
            fs = FileSystemStorage(location=storage)
            file_m = fs.save(f.name,f)
            listz.append(f.name)
        #print(listz)
        list_len = len(listz)
        print(list_len)
        excel_list = []
        read_months = []
        global read_datetime, meter_no
        for element in listz:
            my_cols = ['A','B','C','D']
            df = pd.read_csv(f'{storage}/{element}',names = my_cols)
            data = df.iat[1, 1]
            # stream = data.replace('-', '')
            meter_nox = data.split()[0]
            meter_no = meter_nox.replace('-', '')
            csv_date = data.split()[1]
            actual_date = csv_date.split('/')
            i = int(actual_date[1])
            our_year = int(actual_date[2])
            our_month = calendar.month_name[i]
            our_file_name = f'{meter_no} - {our_month} {our_year}'
            if not os.path.exists(f'{our_month}_{our_year}_files'):
                os.makedirs(f'{our_month}_{our_year}_files')
            df.to_excel(f'{our_month}_{our_year}_files/{our_file_name}.xlsx',index=None,header=False,sheet_name='sheet1')
            excel_list.append(f'{our_file_name}.xlsx')
        print(excel_list)


        for excel in excel_list:
            #df1 = pd.read_excel(f'{our_month}_files/{excel}')
            #print(df1)
            read_months.append(our_month)
            workbook = py.load_workbook(f"{our_month}_{our_year}_files/{excel}")
            sheet = workbook['sheet1']
            man_value = sheet['B1'].value
            manu = man_value.split()[0]  # meter manufacturer
            # print(manu)
            input_string = str(sheet["B2"].value)
            meter = input_string.split()[0].replace('-', '')  # meter number
            # print(meter)
            read_datetime = f'{input_string.split()[1]} {input_string.split()[2]}'
            date_pattern = '%{MONTHNUM:month}/%{MONTHDAY:day}/%{YEAR:year}'
            grok = Grok(date_pattern)
            xl = (grok.match(input_string))
            mm = xl['month']
            dd = xl['day']
            yr = xl['year']
            x = datetime(int(yr), int(mm), int(dd))
            # ll = (x.strftime("%d-%m-%Y"))
            D_O_R = (x.strftime("%d-%b-%Y"))  # reading date
            list3, list4, list5, list6, list7, list8, list9, list10, list11, list12 = [], [], [], [], [], [], [], [], [], []
            list13, list14, list15, list16, list17, list18, list19, list20, list21, list22 = [], [], [], [], [], [], [], [], [], []
            list23, list24, list25, list26, list27, list28, list29, list30, list31, list32 = [], [], [], [], [], [], [], [], [], []
            list1a, list33, list34 = [], [], []

            for r in range(1, sheet.max_row):
                for x in range(1, sheet.max_column):
                    cell = sheet.cell(r, x)
                    #print(sheet.max_row)
                    if cell.value == 'Cumulative totals':
                        cell_imp_whr = sheet.cell(r + 3, x + 1)
                        cell_unit = sheet.cell(r + 3, x + 2)
                        cell_expwh_unit = sheet.cell(r + 4, x + 2)
                        #imp_wh = ''
                        if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                            imp_whr = cell_imp_whr.value.replace(',', '')
                            imp_whh = round(float(imp_whr) / 1000000, 3)
                            imp_wh = "{:,}".format(imp_whh)
                            # print(float(imp_wh))
                        #exp_wh = ''
                        cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                        if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                            exp_whr = cell_exp_whr.value.replace(',', '')
                            exp_whh = round(float(exp_whr) / 1000000, 3)
                            exp_wh = "{:,}".format(exp_whh)

                        #app_pow = ''
                        cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                        cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                        if cell_app_pow_unit.value == 'VAh':
                            app_powr = cell_app_pow.value.replace(',', '')
                            app_powh = round(float(app_powr) / 1000000, 3)
                            app_pow = "{:,}".format(app_powh)
                            # app_pow = float(app_powr)

                        # getting total import and export MVarh
                        #total_imp_MVArh = ''

                        unit1_MVArh = sheet.cell(r + 5, x + 2)
                        unit3_MVArh = sheet.cell(r + 7, x + 2)
                        if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                            imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                            imp_MVArh1 = float(imp_MVArh1_r)
                            imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                            imp_MVArh2 = float(imp_MVArh2_r)
                            total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                            total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                        #total_exp_MVArh = ''
                        #total_MVArh = ''
                        unit2_MVArh = sheet.cell(r + 6, x + 2)
                        unit4_MVArh = sheet.cell(r + 8, x + 2)
                        if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting import MVArh. We add Q3 & Q4
                            exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                            exp_MVArh1 = float(exp_MVArh1_r)
                            exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                            exp_MVArh2 = float(exp_MVArh2_r)
                            total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                            total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)

                        #total MVArh

                        total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(total_imp_MVArh.replace(',', ''))
                        total_MVArh_xx = round(total_MVArh_x, 3)
                        total_mvarh = "{:,}".format(total_MVArh_xx)

                    #rate1,rate2,rate1,rate3,rate4,rate6 = '','','','','',''
                    if cell.value == 'Register' and sheet.cell(r-2,x).value == 'Rates' and sheet.cell(r+1,x+2).value == 'Import Wh':  # getting the various rates
                        unit1 = sheet.cell(r + 1, x + 2)
                        unit2 = sheet.cell(r + 2, x + 2)
                        unit3 = sheet.cell(r + 3, x + 2)
                        if unit3.value == unit2.value == 'Import Wh':
                            rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                            rate1_rh = round(float(rate1_r) / 1000000, 3)
                            rate1 = "{:,}".format(rate1_rh)

                        if unit1.value == unit3.value == 'Import Wh':
                            rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                            rate2_rh = round(float(rate2_r) / 1000000, 3)
                            rate2 = "{:,}".format(rate2_rh)

                        if unit2.value == unit1.value == 'Import Wh':
                            rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                            rate3_rh = round(float(rate3_r) / 1000000, 3)
                            rate3 = "{:,}".format(rate3_rh)

                        unit4 = sheet.cell(r + 4, x + 2)
                        unit5 = sheet.cell(r + 5, x + 2)
                        unit6 = sheet.cell(r + 6, x + 2)
                        if unit5.value == unit6.value == 'Export Wh':
                            rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                            rate4_rh = round(float(rate4_r) / 1000000, 3)
                            rate4 = "{:,}".format(rate4_rh)

                        if unit4.value == unit6.value == 'Export Wh':
                            rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                            rate5_rh = round(float(rate5_r) / 1000000, 3)
                            rate5 = "{:,}".format(rate5_rh)

                        if unit4.value == unit5.value == 'Export Wh':
                            rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                            rate6_rh = round(float(rate6_r) / 1000000, 3)
                            rate6 = "{:,}".format(rate6_rh)

                    if cell.value == 'Billing event details':
                        req_cell = sheet.cell(r + 2, x)
                        req_cell1 = sheet.cell(r + 3, x)
                        if req_cell.value == 'Billing reset number:':
                            resets = sheet.cell(r + 2, x + 1).value
                            # no_of_resets = float(no_of_resets_r)

                        if req_cell1.value == 'Time of billing reset:':
                            reset_time = sheet.cell(r + 3, x + 1).value
                    #global sales_month

                    #sales_month = 0
                    if cell.value == 'Time of billing reset:' and sheet.cell(r + 1,x).value == 'Billing period end date':
                        #M_O_S_r = sheet.cell(r + 1, x + 1).value
                        M_O_S_r = sheet.cell(r, x + 1).value
                        #print(M_O_S_r)
                        M_O_S_x = M_O_S_r.split('/')
                        M_O_S_i = int(M_O_S_x[1])
                        sales_month = rate5
                        #sales_month = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'


                    #max_dem_1,max_dem_2,max_dem_3 = '','',''
                    pow_down_count, pow_down_dt, prog_count, prog_count_dt = '', '', '', ''
                    if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,x).value == 'Register':  # getting maximum demand
                        unit_1 = sheet.cell(r + 3, x + 2)
                        unit_2 = sheet.cell(r + 4, x + 2)
                        unit_3 = sheet.cell(r + 5, x + 2)
                        if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                            max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                            max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 4)
                            max_dem_1 = "{:,}".format(max_dem_1_rh)

                            max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                            max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 4)
                            max_dem_2 = "{:,}".format(max_dem_2_rh)

                            max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                            max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 4)
                            max_dem_3 = "{:,}".format(max_dem_3_rh)

                    #max_dem_1t, max_dem_2t, max_dem_3t, = '','',''
                    if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                            r + 2,
                            x + 3).value == 'Time and date':  # getting max-demands time
                        unit_1t = sheet.cell(r + 3, x + 3)
                        unit_2t = sheet.cell(r + 4, x + 3)
                        unit_3t = sheet.cell(r + 5, x + 3)

                        max_dem_1t = unit_1t.value
                        max_dem_2t = unit_2t.value
                        max_dem_3t = unit_3t.value

                    if cell.value == 'Billing period start date' and sheet.cell(r+1,x).value == 'Billing period end date':
                        sales_value = sheet.cell(r,x+1).value
                        sales_x = sales_value.split('/')
                        month_int = int(sales_x[1])
                        this_month = calendar.month_name[month_int]
                        sales = f'{this_month} {sales_x[2]}'






            c.execute(f"""CREATE TABLE IF NOT EXISTS {meter} (id int(5) NOT NULL AUTO_INCREMENT PRIMARY KEY,time_inserted VARCHAR(20),
            inserted_by varchar(10),meter_read_by varchar(10),reading_datetime varchar(20),meter_no varchar(10),Energy_For varchar(20),
            Cum_Import VARCHAR(13),Cum_Export VARCHAR(13),Apparent_Power VARCHAR(13),Rate_1 VARCHAR(13),Rate_2 VARCHAR(13),
            Rate_3 VARCHAR(13),Rate_4 VARCHAR(13),Rate_5 VARCHAR(13),Rate_6 VARCHAR(13),Max_Dem_1 VARCHAR(13),Max_Dem1_time VARCHAR(13),
            Max_Dem_2 VARCHAR(13),Max_Dem2_time VARCHAR(13),Max_Dem_3 VARCHAR(13),Max_Dem3_time VARCHAR(13),Import_MVArh VARCHAR(13),
            Export_MVArh VARCHAR(13),Total_MVAh VARCHAR(13),No_of_Resets int(5),Last_Reset varchar(15),Power_Down_Count varchar(5),
            Lst_pwr_dwn_date_and_time varchar(20),prog_count int(5),last_prog_date varchar(20))""")

            sql = f"""INSERT INTO {meter} (time_inserted,inserted_by,meter_read_by,reading_datetime,meter_no, Energy_For, Cum_Import, Cum_Export, Apparent_Power,Rate_1,
            Rate_2, Rate_3, Rate_4, Rate_5, Rate_6, Max_Dem_1, Max_Dem1_time, Max_Dem_2,Max_Dem2_time, Max_Dem_3, Max_Dem3_time, Import_MVArh,
            Export_MVArh,Total_MVAh,No_of_Resets,Last_Reset,Power_Down_Count,Lst_pwr_dwn_date_and_time,prog_count,last_prog_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            val = (unix, active_user, active_user, read_datetime, meter,sales, imp_wh, exp_wh, app_pow, rate1, rate2,
            rate3, rate4, rate5,rate6, max_dem_1, max_dem_1t, max_dem_2, max_dem_2t, max_dem_3, max_dem_3t, total_imp_MVArh, total_exp_MVArh,
            total_mvarh, resets, reset_time, pow_down_count, pow_down_dt, prog_count, prog_count_dt)

            c.execute(sql,val)
            conn.commit()

    return render(request, 'form_success.html')
    #c.close()
    conn.close()










        
        









def submission(request): #iNSERTING NONTHLY RECORD
    if request.method == 'POST':
        global active_user
        unix = datetime.now().replace (microsecond=0)
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        #substation = request.POST['subst']
        M_O_S = request.POST['month_of_sale']
        R_D = request.POST['reading_date']
        R_T = request.POST['reading_time']
        meter_no = request.POST['main_meter_no']
        cum_imp = request.POST['cum_imp_m'].replace(',','')
        cum_exp = request.POST['cum_exp_m'].replace(',','')
        app_pow = request.POST['app_pow_m'].replace(',','')
        rate1 = request.POST['rate1_m'].replace(',','')
        rate2 = request.POST['rate2_m'].replace(',','')
        rate3 = request.POST['rate3_m'].replace(',','')
        rate4 = request.POST['rate4_m'].replace(',','')
        rate5 = request.POST['rate5_m'].replace(',','')
        rate6 = request.POST['rate6_m'].replace(',','')
        max_dem1 = request.POST['max_dem_m']
        max_dem1_time = request.POST['max_dem_t_m']
        max_dem2 = request.POST['max_dem2_m']
        max_dem2_time = request.POST['max_dem2_t_m']
        max_dem3 = request.POST['max_dem3_m']
        max_dem3_time = request.POST['max_dem3_t_m']
        imp_mvarh = request.POST['imp_mvarh_m'].replace(',','')
        exp_mvarh = request.POST['exp_mvarh_m'].replace(',','')
        total_mvarh = request.POST['total_mvarh_m'].replace(',','')
        resets = request.POST['resets_m']
        last_reset = request.POST['last_reset_m']
        pow_down_count = request.POST['pow_down_count_m']
        pow_down_dt = request.POST['pow_down_dt_m']
        prog_count = request.POST['prog_count_m']
        prog_count_dt = request.POST['prog_count_dt_m']
        vt_ratio = request.POST['vt_ratio_m']
        ct_ratio = request.POST['ct_ratio_m']

        #date for check meter
        meter_no_ch = request.POST['check_meter_no']
        cum_imp_ch = request.POST['cum_imp_c'].replace(',', '')
        cum_exp_ch = request.POST['cum_exp_c'].replace(',', '')
        app_pow_ch = request.POST['app_pow_c'].replace(',', '')
        rate1_ch = request.POST['rate1_c'].replace(',', '')
        rate2_ch = request.POST['rate2_c'].replace(',', '')
        rate3_ch = request.POST['rate3_c'].replace(',', '')
        rate4_ch = request.POST['rate4_c'].replace(',', '')
        rate5_ch = request.POST['rate5_c'].replace(',', '')
        rate6_ch = request.POST['rate6_c'].replace(',', '')
        max_dem1_ch = request.POST['max_dem_c']
        max_dem1_time_ch = request.POST['max_dem_t_c']
        max_dem2_ch = request.POST['max_dem2_c']
        max_dem2_time_ch = request.POST['max_dem2_t_c']
        max_dem3_ch = request.POST['max_dem3_c']
        max_dem3_time_ch = request.POST['max_dem3_t_c']
        imp_mvarh_ch = request.POST['imp_mvarh_c'].replace(',', '')
        exp_mvarh_ch = request.POST['exp_mvarh_c'].replace(',', '')
        total_mvarh_ch = request.POST['total_mvarh_c'].replace(',', '')
        resets_ch = request.POST['resets_c']
        last_reset_ch = request.POST['last_reset_c']
        pow_down_count_ch = request.POST['pow_down_count_c']
        pow_down_dt_ch = request.POST['pow_down_dt_c']
        prog_count_ch = request.POST['prog_count_c']
        prog_count_dt_ch = request.POST['prog_count_dt_c']
        vt_ratio_ch = request.POST['vt_ratio_c']
        ct_ratio_ch = request.POST['ct_ratio_c']
        #current_user = request.user

        #active_user = user_act
        #active_user = read_by
        #inserted_by = active_user
        read_time= request.POST['reading_time']
        read_date = request.POST['reading_date']
        #id = 0

        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        sql = """INSERT INTO lgg_main(time_stamp,inserted_by,meter_read_by,reading_date,reading_time,meter_no, Energy_For, Cum_Import, Cum_Export, Apparent_Power,Rate_1,
        Rate_2, Rate_3, Rate_4, Rate_5, Rate_6, Max_Dem_1, Max_Dem1_time, Max_Dem_2,Max_Dem2_time, Max_Dem3, Max_Dem3_time, Import_MVArh,
        Export_MVArh,Total_MVAh,No_of_Resets,Last_Reset,Power_Down_Count,Lst_pwr_dwn_date_and_time,prog_count,last_prog_date,VT_Ratio,CT_Ratio,
        meter_no_ch,Cum_Import_ch, Cum_Export_ch, Apparent_Power_ch,Rate_1_ch,
        Rate_2_ch, Rate_3_ch, Rate_4_ch, Rate_5_ch, Rate_6_ch, Max_Dem_1_ch, Max_Dem1_time_ch, Max_Dem_2_ch,Max_Dem2_time_ch, Max_Dem3_ch,
         Max_Dem3_time_ch, Import_MVArh_ch,Export_MVArh_ch,Total_MVAh_ch,No_of_Resets_ch,Last_Reset_ch,
        Power_Down_Count_ch,Lst_pwr_dwn_date_and_time_ch,prog_count_ch,last_prog_date_ch,VT_Ratio_ch,CT_Ratio_ch)
         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,
         %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)"""
        #sql2 = 'read username FROM auth_user'

        val = (unix, active_user, active_user,read_date,read_time, meter_no, M_O_S, cum_imp, cum_exp, app_pow, rate1, rate2, rate3, rate4, rate5,
               rate6, max_dem1, max_dem1_time, max_dem2, max_dem2_time, max_dem3, max_dem3_time, imp_mvarh, exp_mvarh,
               total_mvarh, resets, last_reset, pow_down_count, pow_down_dt, prog_count, prog_count_dt, vt_ratio, ct_ratio,
               meter_no_ch,cum_imp_ch,cum_exp_ch,app_pow_ch,rate1_ch,rate2_ch,rate3_ch,rate4_ch,rate5_ch,rate6_ch,max_dem1_ch,max_dem1_time_ch,
               max_dem2_ch,max_dem2_time_ch,max_dem3_ch,max_dem3_time_ch,imp_mvarh_ch,exp_mvarh_ch,total_mvarh_ch,resets_ch,last_reset_ch,
               pow_down_count_ch,pow_down_dt_ch,prog_count_ch,prog_count_dt_ch,vt_ratio_ch,ct_ratio_ch)

        c.execute(sql,val)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')
#global df1,df2,df3,df_ia,df_ib,df_ic,df_va,df_vb,df_vc,df_pw,df_vars
def monthly_LP(request): # Inserting Load Profile
    engine = create_engine('mysql://admin:556a6069d688f13dd7a9eb566a5fef92588605cf57e7bab5@165.227.253.239:3306/meteringdatabase')
    if request.method == 'POST':
        global active_user
        uploaded_file = request.FILES['LP']
        fs = FileSystemStorage()
        file_LP = fs.save(uploaded_file.name, uploaded_file)

        now = datetime.now()
        xlm = f"Meter Load Profile For {now.strftime('%B')}"
        fd = pd.read_csv(f'media/{file_LP}')
        fd.to_excel(f"loadprofiles/{xlm}.xlsx", index=False)
        #month = now.strftime('%B')
        start = time.time()
        #workbook = xlsxwriter.Workbook('Example1.xlsx')
        #worksheet = workbook.add_worksheet()
        wb = xl.load_workbook(f"loadprofiles/{xlm}.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        input_string = str(ws["B2"].value)
        # print(input_string)
        # global meter
        meter_no = input_string.split()[0].replace('-', '')  # meter number
        for i in range(1, ws.max_column):
            cell_value = ws.cell(row=1, column=i).value
            # if cell_value1 != '':
            if not cell_value == None:
                # cell_value = cell_value1.replace('-','')
                manu_1 = 'Elster'
                if cell_value.find(manu_1) == 1:
                    print(f'manufacturer is Elster')
                    break

        meter_no_x = ws["B2"].value
        meter_no = meter_no_x.split()[0].replace('-', '')
        #print(meter_no)
        # engine = create_engine('mysql://root:@localhost/meteringdatabase')
        lista = []
        list_1 = []
        list_2 = []
        list_3 = []
        list_4 = []
        for x in range(1, 20):
            for r in range(1, ws.max_column + 1):
                cell = ws.cell(r, x)
                global df,df1,df2,df3,df4
                if cell.value == 'Date' and ws.cell(r, x + 1).value == 'Start' and ws.cell(r,
                                                                                           x + 2).value == 'End':  # check this loop for x and r (row and column)
                    for m in range(r + 1, ws.max_row+1):
                        use_date = ws.cell(m, x).value
                        date_time_obj = datetime.strptime(use_date, '%d/%m/%Y').date()
                        #print(date_time_obj)
                        month = date_time_obj.strftime("%B")
                        #month = use_date.strftime('%B')
                        start_time_1 = ws.cell(m, x + 1).value
                        start_time = '%.6s' % start_time_1
                        end_time_1 = ws.cell(m, x + 2).value
                        end_time = '%.6s' % end_time_1
                        list_1.append(use_date)
                        list_2.append(start_time)
                        list_3.append(end_time)
                        list_4.append(month)
                    # print(list_1)
                    df1 = pd.DataFrame(list_1)
                    df1.columns = ['date']
                    # print(df1)
                    df2 = pd.DataFrame(list_2)
                    df2.columns = ['start_time']
                    # print(df2)
                    df3 = pd.DataFrame(list_3)
                    df3.columns = ['end_time']
                    df4 = pd.DataFrame(list_4)
                    df4.columns = ['month']
                    # print(df3)
                    #df = pd.concat([df1, df2, df3], axis=1)

                    # print(df)

                list_ia = []
                # row_ia = 0
                global dfnew
                global df_ia
                if cell.value == 'A: PhA: Av':
                    for m in range(r, ws.max_row+1):
                        I_A = ws.cell(m, x - 1).value
                        list_ia.append(I_A)
                        lista.append(active_user)
                    df_ia = pd.DataFrame(list_ia)
                    df_ia.columns = ['I_A']
                    df_au = pd.DataFrame(lista)
                    df_au.columns = ['inserted_by']

                    #dfnew = pd.concat([df, df_ia], axis=1)
                    # print(dfnew)

                    break

                list_ib = []
                global df_ib
                if cell.value == 'A: PhB: Av':
                    for m in range(r, ws.max_row+1):
                        I_B = ws.cell(m, x - 1).value
                        list_ib.append(I_B)
                    df_ib = pd.DataFrame(list_ib)
                    df_ib.columns = ['I_B']
                    #df_ib = pd.concat([dfnew, df_ib_x], axis=1)
                    # print(df_ib)
                    break

                list_ic = []
                global df_ic
                if cell.value == 'A: PhC: Av':
                    for m in range(r, ws.max_row+1):
                        I_C = ws.cell(m, x - 1).value
                        list_ic.append(I_C)
                    df_ic = pd.DataFrame(list_ic)
                    df_ic.columns = ['I_C']
                    #df_ic = pd.concat([df_ib, df_ic_x], axis=1)
                    #print(df_ic)
                    break

                list_va = []
                global df_va
                if cell.value == 'V: PhA: Av':
                    for m in range(r, ws.max_row+1):
                        V_A = ws.cell(m, x - 1).value
                        list_va.append(V_A)
                    df_va = pd.DataFrame(list_va)
                    df_va.columns = ['V_A']
                    break

                global df_vb
                list_vb = []
                if cell.value == 'V: PhB: Av':
                    for m in range(r, ws.max_row+1):
                        V_B = ws.cell(m, x - 1).value
                        list_vb.append(V_B)
                    df_vb = pd.DataFrame(list_vb)
                    df_vb.columns = ['V_B']
                    break
                global df_vc
                list_vc = []
                if cell.value == 'V: PhC: Av':
                    for m in range(r, ws.max_row+1):
                        V_C = ws.cell(m, x - 1).value
                        list_vc.append(V_C)
                    df_vc = pd.DataFrame(list_vc)
                    df_vc.columns = ['V_C']
                    break

                global df_pw
                list_pw = []
                if cell.value == 'kW: Sys: Av':
                    for m in range(r, ws.max_row+1):
                        PW = ws.cell(m, x - 1).value
                        # print(PW)
                        list_pw.append(PW)
                    df_pw = pd.DataFrame(list_pw)
                    df_pw.columns = ['PW']
                    break

                global df_Pvar
                list_Pvar = []
                if cell.value == 'kvar: Sys: Av':
                    for m in range(r, ws.max_row+1):
                        Pvar = ws.cell(m, x - 1).value
                        # print(PW)
                        list_Pvar.append(Pvar)
                    df_Pvar = pd.DataFrame(list_Pvar)
                    df_Pvar.columns = ['P_var']
                    #now = datetime.datetime.now().date()


                    df = pd.concat([df_au,df1,df4,df2,df3,df_ia,df_ib,df_ic,df_va,df_vb,df_vc,df_pw,df_Pvar],axis=1)
                    #print(df)
                    df.to_sql(f'{meter_no}_Load_Profile', con=engine, if_exists='append', index=False,dtype={'inserted_by': VARCHAR(length=10),'date': VARCHAR(length=30),'month': VARCHAR(length=12),
                                        'start_time': String(length=10), 'end_time': String(length=10),
                                        'I_A': String(length=10), 'I_B': String(length=10),
                                        'I_C': String(length=10),'V_A':VARCHAR(length=10),'V_B':VARCHAR(length=10),'V_C':VARCHAR(length=10),
                                                        'PW':VARCHAR(length=10),'P_var':VARCHAR(length=10),})


                    break

        end = time.time()
        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        c.execute(f"ALTER TABLE {meter_no}_Load_Profile ADD COLUMN IF NOT EXISTS `id` int(10) UNSIGNED PRIMARY KEY AUTO_INCREMENT FIRST")
        conn.commit()
        c.close()
        conn.close()
        print("Elapsed time is  {}".format(end - start))

        return render(request,'form_success.html')

def LP_plot(request): #Plotting Load profiles
    conn = mysql.connector.connect(**db_config)
    c = conn.cursor()
    sql = "SELECT start_time, PW FROM uetcl0103 WHERE month = 'September' AND date = '22/09/2020'"
    time = []
    power = []
    c.execute(sql)
    data = c.fetchall()
    for row in data:
        time.append(row[0])
        power.append(float(row[1])/1000)

    # print(power)
    plt.rcParams.update({'font.size': 22})
    fig = plt.figure(figsize=(15, 6))
    plt.ylabel("Power/kW")
    plt.xlabel('Time/Hrs')
    plt.plot(time, power, color='red', )
    plt.title("Load Profile For January")
    #plt.show()
    mpld3.save_html(fig, 'E:\\METERING DATABASE\\metering_database\\templates\\kim.html')

    conn.commit()
    c.close()
    conn.close()

    return render(request, 'kim.html') #end of function



def umeme(request):
    #global meters_umeme
    meters_umeme = []
    conn = mysql.connector.connect(**db_config)
    c = conn.cursor()
    sql = "SELECT meter_no FROM substation_meters WHERE distributor = 'umeme'"
    c.execute(sql)
    data = c.fetchall()
    for row in data:
        meters_umeme.append(row[0])
    # print(meters)
    conn.commit()
    c.close()
    conn.close()
    year = date.today().year
    xxl = date.today().month - 1
    # print(now)
    month = calendar.month_name[xxl]
    bill_date = f'{month} {year}'

    #{"distributor": distributor,}
    #months = []
    #for name in calendar.month_name:
        # print(name)
        #months.append(name)
    #years = []
    #for i in range(-2, 5):
        #year = date.today().year + i
        #years.append(year)
    #print(years)

    #meter = ['UETCL001', 'UETCL002', 'UETCL003','UETCL004']
    #distributors = ['Umeme','UEDCL','PACKMECS','BEL','KRECS']
    #print(list1)
    return render(request, 'gen_bill.html',{'meters_umeme':meters_umeme,'bill_date':bill_date})

def bill_gen(request):
    if request.method == 'POST':
        meter_no = request.POST['meter_no']
        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        sql = "SELECT reading_date, reading_time, meter_no_ch,"
        conn.commit()
        c.close()
        conn.close()

    return render(request,'bill.html')

def newmeter(request):
    return render(request,'new_meter_home.html')

def standalone(request):
    return render(request,'standalone_meter.html')

def substation(request):
    return render(request,'substation_meter.html')



def querries(request):
    return render(request, 'querries.html')

def querriesback(request):
    return render(request, 'welcome.html')

def new_LP(request):
    return render(request,'insertLP.html')

def hist_records(request): #inserting historical data
    if request.method == 'POST':
        global active_user
        uploaded_file = request.FILES['hist']
        fs = FileSystemStorage()
        file_m = fs.save(uploaded_file.name, uploaded_file)
        wb = Workbook()
        ws = wb.active
        with open(f'media/{file_m}', 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(f'historical_records/{file_m}.xlsx')
        workbook = py.load_workbook(f'historical_records/{file_m}.xlsx')
        sheet = workbook['Sheet']
        man_value = sheet['B1'].value
        manu = man_value.split()[0]  # meter manufacturer
        # print(manu)
        input_string = str(sheet["B2"].value)
        # print(input_string)
        # global meter
        meter_no = input_string.split()[0].replace('-', '')  # meter number
        # print(meter)
        read_time = f'{input_string.split()[1]} {input_string.split()[2]}'  # reading time

        # global meter_no
        # sql_1 = "SELECT CTratio, VTratio FROM meter_details WHERE meter_no=?",((meter,))
        # c.execute("SELECT CTratios, VTratio FROM standalone_meter_details WHERE meter_no = %s", meter)
        # data = c.fetchone()

        date_pattern = '%{MONTHNUM:month}/%{MONTHDAY:day}/%{YEAR:year}'
        grok = Grok(date_pattern)
        xl = (grok.match(input_string))
        mm = xl['month']
        dd = xl['day']
        yr = xl['year']
        x = datetime(int(yr), int(mm), int(dd))
        # ll = (x.strftime("%d-%m-%Y"))
        D_O_R = (x.strftime("%d-%b-%Y"))  # reading date
        list1 = []
        list2 = []
        list7 = []
        for m in range(1, sheet.max_row + 1):
            for n in range(1, sheet.max_column + 1):
                # cell1 = sheet.cell(m,n)
                cell = sheet.cell(m, n).value
                cell_value = str(cell)
                if (cell_value.find('Historical data set') != -1):
                    list1.append(cell_value)
                    list2.append(m)
        print(len(list1))
        print(list2)
        list3 = []
        list4 = []
        list5 = []
        list6 = []
        list7 = []
        list8 = []
        list9 = []
        list10 = []
        list11 = []
        list12 = []
        list13 = []
        list14 = []
        list15 = []
        list16 = []
        list17 = []
        list18 = []
        list19 = []
        list20 = []
        list21 = []
        list22 = []
        list23 = []
        list24 = []
        list25 = []
        list26 = []
        list27 = []
        list28 = []
        list29 = []
        list30 = []
        list31 = []
        list32 = []
        list1a = []
        list33 = []
        list34 = []

        for i in range(1, len(list2) + 1):
            if i <= len(list2) - 1:
                row_1 = list2[i - 1]
                row_2 = list2[i]
                # print(row_2)
                for r in range(row_1, row_2 + 1):
                    for x in range(1, sheet.max_column + 1):
                        cell = sheet.cell(r, x)
                        # global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                        if cell.value == 'Cumulative totals':
                            cell_imp_whr = sheet.cell(r + 3, x + 1)
                            cell_unit = sheet.cell(r + 3, x + 2)
                            cell_expwh_unit = sheet.cell(r + 4, x + 2)
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                                imp_whr = cell_imp_whr.value.replace(',', '')
                                imp_whh = round(float(imp_whr) / 1000000, 3)
                                imp_wh = "{:,}".format(imp_whh)
                                list1a.append(imp_wh)

                            cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                                exp_whr = cell_exp_whr.value.replace(',', '')
                                exp_whh = round(float(exp_whr) / 1000000, 3)
                                exp_wh = "{:,}".format(exp_whh)
                                list4.append(exp_wh)

                            cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                            cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                            if cell_app_pow_unit.value == 'VAh':
                                app_powr = cell_app_pow.value.replace(',', '')
                                app_powh = round(float(app_powr) / 1000000, 3)
                                app_pow = "{:,}".format(app_powh)
                                list5.append(app_pow)
                                # app_pow = float(app_powr)

                            # getting total import and export MVarh
                            unit1_MVArh = sheet.cell(r + 5, x + 2)
                            unit3_MVArh = sheet.cell(r + 7, x + 2)
                            global total_imp_MVArh
                            if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                                imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                imp_MVArh1 = float(imp_MVArh1_r)
                                imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                                imp_MVArh2 = float(imp_MVArh2_r)
                                total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                                total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                                list6.append(total_imp_MVArh)

                            unit2_MVArh = sheet.cell(r + 6, x + 2)
                            unit4_MVArh = sheet.cell(r + 8, x + 2)
                            if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                                exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                exp_MVArh1 = float(exp_MVArh1_r)
                                exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                                exp_MVArh2 = float(exp_MVArh2_r)
                                total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                                total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                                list7.append(total_exp_MVArh)

                                # total MVArh
                                total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(
                                    total_imp_MVArh.replace(',', ''))
                                total_MVArh_xx = round(total_MVArh_x, 3)
                                total_MVArh = "{:,}".format(total_MVArh_xx)
                                list8.append(total_MVArh)

                        # global rate1,rate2,rate3,rate4,rate5,rate6
                        if cell.value == 'Register':  # getting the various rates
                            unit1 = sheet.cell(r + 1, x + 2)
                            unit2 = sheet.cell(r + 2, x + 2)
                            unit3 = sheet.cell(r + 3, x + 2)
                            if unit3.value == unit2.value == 'Import Wh':
                                rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                                rate1_rh = round(float(rate1_r) / 1000000, 3)
                                rate1 = "{:,}".format(rate1_rh)
                                list9.append(rate1)

                            if unit1.value == unit3.value == 'Import Wh':
                                rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                                rate2_rh = round(float(rate2_r) / 1000000, 3)
                                rate2 = "{:,}".format(rate2_rh)
                                list10.append(rate2)

                            if unit2.value == unit1.value == 'Import Wh':
                                rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                                rate3_rh = round(float(rate3_r) / 1000000, 3)
                                rate3 = "{:,}".format(rate3_rh)
                                list11.append(rate3)

                            unit4 = sheet.cell(r + 4, x + 2)
                            unit5 = sheet.cell(r + 5, x + 2)
                            unit6 = sheet.cell(r + 6, x + 2)
                            if unit5.value == unit6.value == 'Export Wh':
                                rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                                rate4_rh = round(float(rate4_r) / 1000000, 3)
                                rate4 = "{:,}".format(rate4_rh)
                                list12.append(rate4)

                            if unit4.value == unit6.value == 'Export Wh':
                                rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                rate5_rh = round(float(rate5_r) / 1000000, 3)
                                rate5 = "{:,}".format(rate5_rh)
                                list13.append(rate5)
                            # global rate6
                            if unit4.value == unit5.value == 'Export Wh':
                                rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                rate6_rh = round(float(rate6_r) / 1000000, 3)
                                rate6 = "{:,}".format(rate6_rh)
                                list14.append(rate6)
                        # global reset_time, resets
                        if cell.value == 'Billing event details':
                            req_cell = sheet.cell(r + 2, x)
                            req_cell1 = sheet.cell(r + 3, x)
                            if req_cell.value == 'Billing reset number:':
                                resets = sheet.cell(r + 2, x + 1).value
                                # no_of_resets = float(no_of_resets_r)
                                list23.append(resets)

                            if req_cell1.value == 'Time of billing reset:':
                                reset_time = sheet.cell(r + 3, x + 1).value
                                list24.append(reset_time)
                        # global M_O_S,M_O_S_r
                        if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                                    x).value == 'Billing period end date':
                            M_O_S_r = sheet.cell(r, x + 1).value
                            M_O_S_x = M_O_S_r.split('/')
                            M_O_S_i = int(M_O_S_x[1])
                            M_O_S = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'
                            list16.append(M_O_S)
                        # global max_dem_1,max_dem_2,max_dem_3
                        if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                                     x).value == 'Register':  # getting maximum demands

                            unit_1 = sheet.cell(r + 3, x + 2)
                            unit_2 = sheet.cell(r + 4, x + 2)
                            unit_3 = sheet.cell(r + 5, x + 2)
                            if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                                max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                                max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                                max_dem_1 = "{:,}".format(max_dem_1_rh)
                                list17.append(max_dem_1)

                                max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                                max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                                max_dem_2 = "{:,}".format(max_dem_2_rh)
                                list18.append(max_dem_2)

                                max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                                max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                                max_dem_3 = "{:,}".format(max_dem_3_rh)
                                list19.append(max_dem_3)
                        # global max_dem_1t, max_dem_2t, max_dem_3t

                        if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                                r + 2,
                                x + 3).value == 'Time and date':  # getting max-demands time
                            unit_1t = sheet.cell(r + 3, x + 3)
                            unit_2t = sheet.cell(r + 4, x + 3)
                            unit_3t = sheet.cell(r + 5, x + 3)

                            max_dem_1t = unit_1t.value
                            max_dem_2t = unit_2t.value
                            max_dem_3t = unit_3t.value
                            list20.append(max_dem_1t)
                            list21.append(max_dem_2t)
                            list22.append(max_dem_3t)

                            # constants
                            unix = datetime.now().replace(second=0, microsecond=0)
                            inserted_by = active_user
                            #active_user = 'kimera'
                            pow_down_count = 'No Data'
                            pow_down_dt = 'No Data'
                            prog_count = 'No Data'
                            prog_count_dt = 'No Data'
                            # list25 = ['No data']
                            list25.append(pow_down_count)
                            list26.append(pow_down_dt)
                            list27.append(prog_count)
                            list28.append(prog_count_dt)
                            list29.append(unix)
                            list30.append(inserted_by)
                            list31.append(meter_no)
                            list32.append(D_O_R)
                            list33.append(read_time)

                i += 1
            elif i == len(list2):
                row_1 = list2[i - 1]
                for r in range(row_1, sheet.max_row):
                    for x in range(1, sheet.max_column + 1):
                        cell = sheet.cell(r, x)
                        # global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                        if cell.value == 'Cumulative totals':
                            cell_imp_whr = sheet.cell(r + 3, x + 1)
                            cell_unit = sheet.cell(r + 3, x + 2)
                            cell_expwh_unit = sheet.cell(r + 4, x + 2)
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                                imp_whr = cell_imp_whr.value.replace(',', '')
                                imp_whh = round(float(imp_whr) / 1000000, 3)
                                imp_wh = "{:,}".format(imp_whh)
                                list1a.append(imp_wh)
                            cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                                exp_whr = cell_exp_whr.value.replace(',', '')
                                exp_whh = round(float(exp_whr) / 1000000, 3)
                                exp_wh = "{:,}".format(exp_whh)
                                list4.append(exp_wh)

                            cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                            cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                            if cell_app_pow_unit.value == 'VAh':
                                app_powr = cell_app_pow.value.replace(',', '')
                                app_powh = round(float(app_powr) / 1000000, 3)
                                app_pow = "{:,}".format(app_powh)
                                list5.append(app_pow)
                                # app_pow = float(app_powr)

                            # getting total import and export MVarh
                            unit1_MVArh = sheet.cell(r + 5, x + 2)
                            unit3_MVArh = sheet.cell(r + 7, x + 2)
                            # global total_imp_MVArh
                            if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                                imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                imp_MVArh1 = float(imp_MVArh1_r)
                                imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                                imp_MVArh2 = float(imp_MVArh2_r)
                                total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                                total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                                list6.append(total_imp_MVArh)

                            unit2_MVArh = sheet.cell(r + 6, x + 2)
                            unit4_MVArh = sheet.cell(r + 8, x + 2)
                            if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                                exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                exp_MVArh1 = float(exp_MVArh1_r)
                                exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                                exp_MVArh2 = float(exp_MVArh2_r)
                                total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                                total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                                list7.append(total_exp_MVArh)

                                # total MVArh
                                total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(
                                    total_imp_MVArh.replace(',', ''))
                                total_MVArh_xx = round(total_MVArh_x, 3)
                                total_MVArh = "{:,}".format(total_MVArh_xx)
                                list8.append(total_MVArh)

                            # global rate1,rate2,rate3,rate4,rate5,rate6
                        if cell.value == 'Register':  # getting the various rates
                            unit1 = sheet.cell(r + 1, x + 2)
                            unit2 = sheet.cell(r + 2, x + 2)
                            unit3 = sheet.cell(r + 3, x + 2)
                            if unit3.value == unit2.value == 'Import Wh':
                                rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                                rate1_rh = round(float(rate1_r) / 1000000, 3)
                                rate1 = "{:,}".format(rate1_rh)
                                list9.append(rate1)

                            if unit1.value == unit3.value == 'Import Wh':
                                rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                                rate2_rh = round(float(rate2_r) / 1000000, 3)
                                rate2 = "{:,}".format(rate2_rh)
                                list10.append(rate2)

                            if unit2.value == unit1.value == 'Import Wh':
                                rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                                rate3_rh = round(float(rate3_r) / 1000000, 3)
                                rate3 = "{:,}".format(rate3_rh)
                                list11.append(rate3)

                            unit4 = sheet.cell(r + 4, x + 2)
                            unit5 = sheet.cell(r + 5, x + 2)
                            unit6 = sheet.cell(r + 6, x + 2)
                            if unit5.value == unit6.value == 'Export Wh':
                                rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                                rate4_rh = round(float(rate4_r) / 1000000, 3)
                                rate4 = "{:,}".format(rate4_rh)
                                list12.append(rate4)

                            if unit4.value == unit6.value == 'Export Wh':
                                rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                rate5_rh = round(float(rate5_r) / 1000000, 3)
                                rate5 = "{:,}".format(rate5_rh)
                                list13.append(rate5)
                            # global rate6
                            if unit4.value == unit5.value == 'Export Wh':
                                rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                rate6_rh = round(float(rate6_r) / 1000000, 3)
                                rate6 = "{:,}".format(rate6_rh)
                                list14.append(rate6)
                            # global reset_time, resets
                        if cell.value == 'Billing event details':
                            req_cell = sheet.cell(r + 2, x)
                            req_cell1 = sheet.cell(r + 3, x)
                            if req_cell.value == 'Billing reset number:':
                                resets = sheet.cell(r + 2, x + 1).value
                                # no_of_resets = float(no_of_resets_r)
                                list23.append(resets)

                            if req_cell1.value == 'Time of billing reset:':
                                reset_time = sheet.cell(r + 3, x + 1).value
                                list24.append(reset_time)
                            # global M_O_S,M_O_S_r
                        if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                                    x).value == 'Billing period end date':
                            M_O_S_r = sheet.cell(r, x + 1).value
                            M_O_S_x = M_O_S_r.split('/')
                            M_O_S_i = int(M_O_S_x[1])
                            M_O_S = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'
                            list16.append(M_O_S)
                            # global max_dem_1,max_dem_2,max_dem_3
                        if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                                     x).value == 'Register':  # getting maximum demands

                            unit_1 = sheet.cell(r + 3, x + 2)
                            unit_2 = sheet.cell(r + 4, x + 2)
                            unit_3 = sheet.cell(r + 5, x + 2)
                            if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                                max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                                max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                                max_dem_1 = "{:,}".format(max_dem_1_rh)
                                list17.append(max_dem_1)

                                max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                                max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                                max_dem_2 = "{:,}".format(max_dem_2_rh)
                                list18.append(max_dem_2)

                                max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                                max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                                max_dem_3 = "{:,}".format(max_dem_3_rh)
                                list19.append(max_dem_3)
                            # global max_dem_1t, max_dem_2t, max_dem_3t

                        if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                                r + 2,
                                x + 3).value == 'Time and date':  # getting max-demands time
                            unit_1t = sheet.cell(r + 3, x + 3)
                            unit_2t = sheet.cell(r + 4, x + 3)
                            unit_3t = sheet.cell(r + 5, x + 3)

                            max_dem_1t = unit_1t.value
                            max_dem_2t = unit_2t.value
                            max_dem_3t = unit_3t.value
                            list20.append(max_dem_1t)
                            list21.append(max_dem_2t)
                            list22.append(max_dem_3t)

                            # constants
                            unix = datetime.now().replace(second=0, microsecond=0)
                            inserted_by = active_user
                            #active_user = 'kimera'
                            pow_down_count = 'No Data'
                            pow_down_dt = 'No Data'
                            prog_count = 'No Data'
                            prog_count_dt = 'No Data'
                            # list25 = ['No data']
                            list25.append(pow_down_count)
                            list26.append(pow_down_dt)
                            list27.append(prog_count)
                            list28.append(prog_count_dt)
                            list29.append(unix)
                            list30.append(inserted_by)
                            list31.append(meter_no)
                            list32.append(D_O_R)
                            list33.append(read_time)

        df1_a = pd.DataFrame(list29)
        df1_a.columns = ['time_stamp']
        df1_b = pd.DataFrame(list30)
        df1_b.columns = ['inserted_by']
        df1_c = pd.DataFrame(list30)
        df1_c.columns = ['meter_read_by']
        df1_e = pd.DataFrame(list33)
        df1_e.columns = ['reading_datetime']
        df1_f = pd.DataFrame(list31)
        df1_f.columns = ['meter_no']
        df1 = pd.DataFrame(list16)
        df1.columns = ['Energy_For']
        df2 = pd.DataFrame(list1a)
        df2.columns = ['Cum_Import']
        df3 = pd.DataFrame(list4)
        df3.columns = ['Cum_Export']
        df4 = pd.DataFrame(list5)
        df4.columns = ['Apparent_Power']
        df5 = pd.DataFrame(list9)
        df5.columns = ['Rate_1']
        df6 = pd.DataFrame(list10)
        df6.columns = ['Rate_2']
        df7 = pd.DataFrame(list11)
        df7.columns = ['Rate_3']
        df8 = pd.DataFrame(list12)
        df8.columns = ['Rate_4']
        df9 = pd.DataFrame(list13)
        df9.columns = ['Rate_5']
        df10 = pd.DataFrame(list14)
        df10.columns = ['Rate_6']
        df11 = pd.DataFrame(list17)
        df11.columns = ['Max_Dem_1']
        df12 = pd.DataFrame(list20)
        df12.columns = ['Max_Dem1_time']
        df13 = pd.DataFrame(list18)
        df13.columns = ['Max_Dem_2']
        df14 = pd.DataFrame(list21)
        df14.columns = ['Max_Dem2_time']
        df15 = pd.DataFrame(list19)
        df15.columns = ['Max_Dem_3']
        df16 = pd.DataFrame(list22)
        df16.columns = ['Max_Dem3_time']
        df17 = pd.DataFrame(list6)
        df17.columns = ['Import_MVArh']
        df18 = pd.DataFrame(list7)
        df18.columns = ['Export_MVArh']
        df19 = pd.DataFrame(list8)
        df19.columns = ['Total_MVAh']
        df20 = pd.DataFrame(list23)
        df20.columns = ['No_of_Resets']
        df21 = pd.DataFrame(list24)
        df21.columns = ['Last_Reset']
        df22 = pd.DataFrame(list25)
        df22.columns = ['Power_Down_Count']
        df23 = pd.DataFrame(list26)
        df23.columns = ['Lst_pwr_dwn_date_and_time']
        df24 = pd.DataFrame(list27)
        df24.columns = ['prog_count']
        df25 = pd.DataFrame(list28)
        df25.columns = ['last_prog_date']

        df = pd.concat(
            [df1_a, df1_b, df1_c, df1_e, df1_f, df1, df2, df3, df4, df5, df6, df7, df8, df9, df10, df11, df12, df13,
             df14, df15, df16,
             df17, df18, df19, df20, df21, df22, df23, df24, df25], axis=1)
        engine = create_engine('mysql://admin:556a6069d688f13dd7a9eb566a5fef92588605cf57e7bab5@165.227.253.239:3306/meteringdatabase')
        df.to_sql(f'{meter_no}_Historical_Data', con=engine, if_exists='append', index=False,
                  dtype={'time_stamp': VARCHAR(length=25), 'inserted_by': VARCHAR(length=20),
                         'meter_read_by': VARCHAR(length=10),
                         'reading_datetime': VARCHAR(length=25), 'meter_no': VARCHAR(length=12),
                         'Energy_For': VARCHAR(length=15), 'Cum_Import': VARCHAR(length=13),
                         'Cum_Export': VARCHAR(length=13), 'Apparent_Power': VARCHAR(length=13),
                         'Rate_1': VARCHAR(length=13), 'Rate_2': VARCHAR(length=13), 'Rate_3': VARCHAR(length=13),
                         'Rate_4': VARCHAR(length=13), 'Rate_5': VARCHAR(length=13), 'Rate_6': VARCHAR(length=13),
                         'Max_Dem_1': VARCHAR(length=13), 'Max_Dem1_time': VARCHAR(length=15),
                         'Max_Dem_2': VARCHAR(length=13), 'Max_Dem2_time': VARCHAR(length=15),
                         'Max_Dem_3': VARCHAR(length=13), 'Max_Dem3_time': VARCHAR(length=15),
                         'Import_MVArh': VARCHAR(length=13), 'Export_MVArh': VARCHAR(length=13),
                         'Total_MVAh': VARCHAR(length=13), 'No_of_Resets': VARCHAR(5),
                         'Last_Reset': VARCHAR(length=25), 'Power_Down_Count': VARCHAR(length=5),
                         'Lst_pwr_dwn_date_and_time': VARCHAR(length=5),
                         'prog_count': VARCHAR(5), 'last_prog_date': VARCHAR(length=25)
                         })
        conn = mysql.connector.connect(**db_config)
        c = conn.cursor()
        c.execute(
            f"ALTER TABLE {meter_no}_Historical_Data ADD COLUMN IF NOT EXISTS `id` int(10) UNSIGNED PRIMARY KEY AUTO_INCREMENT FIRST")
        conn.commit()
        c.close()
        conn.close()


    return render(request, 'form_success.html')
